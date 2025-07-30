import streamlit as st
import pandas as pd
import openpyxl

# ------------------------------
# WRR processing logic
# ------------------------------
def process_wrr_workbook(uploaded_file):
    quarter_label = "3Q25"
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    available_sheets = wb.sheetnames

    def get_sheet_name(base):
        matches = [s for s in available_sheets if base in s and quarter_label in s]
        if not matches:
            raise ValueError(f"❌ Sheet with '{quarter_label} - {base}' not found.")
        return matches[0]

    cell_mappings = [
        ("Max", "Streaming", "F17", "AE17", "AF17", "AN17", "AV17", "AS17", "AJ17", "AM17"),
        ("D+", "Streaming", "F11", "AE11", "AF11", "AN11", "AV11", "AS11", "AJ11", "AM11"),
        ("News", "Consolidated", "F14", "AF14", "AG14", "AO14", "AW14", "AT14", "AK14", "AN11"),
        ("Entertainment", "USN Ent", "F36", "AE36", "AF36", "AN36", "AV36", "AS36", "AJ36", "AM36"),
        ("Sports", "Sports", "F20", "AE20", "AF20", "AN20", "AV20", "AS20", "AJ20", "AM20"),
    ]

    records = []
    for category, sheet_base, fcst, otb, wow, ltg, open_, pipe, hedge, dr_ltg in cell_mappings:
        sheet_name = get_sheet_name(sheet_base)
        ws = wb[sheet_name]
        record = {
            "Quarter": quarter_label,
            "Category": category,
            "Forecast": ws[fcst].value,
            "OTB": ws[otb].value,
            "Booked This Week": ws[wow].value,
            "LTG": ws[ltg].value,
            "Open": ws[open_].value,
            "Pipeline": ws[pipe].value,
            "Hedge": ws[hedge].value,
            "DR LTG": ws[dr_ltg].value,
        }
        records.append(record)

    df = pd.DataFrame(records)

    # # Looker-style table calcs
    # df["OVER_LTG"] = df["Forecast"] - df["LTG"]
    # df["OVER_PIPE"] = df["Forecast"] - df["Pipeline"]
    # df["OVER_OPEN"] = df["Forecast"] - df["Open"]

    return df


# ------------------------------
# Generate Looker table calcs
# ------------------------------
def generate_looker_snippets(df, metrics=["OTB", "LTG", "Pipeline", "Open"]):
    snippets = {}
    for metric in metrics:
        snippet = ""
        for cat in ["Max", "D+", "News", "Entertainment", "Sports"]:
            val = df.loc[df["Category"] == cat, metric].values
            if len(val) > 0:
                value = round(val[0], 2)
                if cat == "Max":
                    snippet += f'if(${{op_staq_sf_union_v.vertical}}="{cat}",{value}+${{booked}}-${{booked}},\n'
                else:
                    snippet += f'if(${{op_staq_sf_union_v.vertical}}="{cat}",{value},\n'
        snippet += "0" + ")" * 5
        snippets[f"OVER_{metric.upper()}"] = snippet
    return snippets


# ------------------------------
# Streamlit App
# ------------------------------
st.set_page_config(page_title="Weekly WRR Report", layout="wide")
st.title("📊 Weekly Digital Sales Report")

uploaded_file = st.file_uploader("Upload Weekly WRR File (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        df = process_wrr_workbook(uploaded_file)

        # Add totals row
        total_row = df.drop(columns=["Category", "Quarter"]).sum(numeric_only=True)
        total_row["Category"] = "Total"
        total_row["Quarter"] = df["Quarter"].iloc[0]
        df_with_total = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        # Format as currency
        currency_cols = df.columns.drop(["Category", "Quarter"])
        styled_df = df_with_total.style.format({col: "${:,.0f}" for col in currency_cols})

        st.subheader("Processed WRR Data")
        st.dataframe(styled_df, use_container_width=True)

        # Looker Snippets
        snippets = generate_looker_snippets(df)
        st.subheader("📄 Looker Table Calcs")
        for name, code in snippets.items():
            st.markdown(f"**{name}**")
            st.code(code, language="sql")

        # Download
        st.download_button(
            label="Download Processed CSV",
            data=df.to_csv(index=False),
            file_name="WRR_Processed.csv",
            mime="text/csv"
        )

    except Exception as e:
        st.error(f"❌ Error: {e}")
else:
    st.info("👈 Upload a WRR Excel file to get started.")
